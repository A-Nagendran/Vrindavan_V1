import streamlit as st
import google.generativeai as genai
import PyPDF2
import pandas as pd
import io
import time
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

# --- Page Config ---
st.set_page_config(
    page_title="Vrindavan Auditor",
    page_icon="🦚",
    layout="wide"
)

# --- Session State Initialization ---
if 'analysis_results' not in st.session_state:
    st.session_state['analysis_results'] = None
if 'csm_summary' not in st.session_state:
    st.session_state['csm_summary'] = None
if 'processing_log' not in st.session_state:
    st.session_state['processing_log'] = []

# --- Helper Functions ---

def extract_text_from_pdf(file_bytes):
    try:
        reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception:
        return ""

def get_gemini_model(api_key):
    try:
        genai.configure(api_key=api_key)
        # Smart Model Selector
        all_models = [m.name for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
        
        # Priority: Flash -> Pro -> Any
        chosen_model = next((m for m in all_models if 'flash' in m.lower() and '1.5' in m), None)
        if not chosen_model:
            chosen_model = next((m for m in all_models if 'pro' in m.lower() and '1.5' in m), None)
        if not chosen_model and all_models:
            chosen_model = all_models[0]
            
        if chosen_model:
            return genai.GenerativeModel(chosen_model)
        return None
    except:
        return None

def analyze_call(model, text):
    prompt = """
    You are a QA Auditor for 'The House of Abhinandan Lodha' (Navratna at Vrindavan).
    Analyze this sales call transcript. Output strictly in English.

    **Extract these 10 Fields:**
    1. Lead Intent (Pure Investor / Spiritual Buyer / Second-home / Mixed)
    2. Brand Trust (Yes/No + Detail)
    3. Value (Yes/No + Detail)
    4. ROI (Yes/No + Detail)
    5. Spiritual (Yes/No + Detail)
    6. Technical (Yes/No + Detail)
    7. Urgency: 6L+2L Offer (Did they use it?)
    8. Urgency: Price Movement (Did they use it?)
    9. Family Objections
    10. Pitch Flow Adherence %

    **OUTPUT FORMAT (Single Line separated by ###):**
    CSM Name###Customer Name###Lead Intent###Brand Trust (Yes/No)###Brand Trust Detail###Value (Yes/No)###Value Detail###ROI (Yes/No)###ROI Detail###Spiritual (Yes/No)###Spiritual Detail###Technical (Yes/No)###Technical Detail###Urgency: Offers###Urgency: Price Move###Family Objections###Pitch Flow %
    
    **TRANSCRIPT:**
    """ + text[:30000]

    try:
        safe = [
            {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"}
        ]
        response = model.generate_content(prompt, safety_settings=safe)
        raw = response.text.strip()
        parts = [x.strip() for x in raw.split('###')]
        while len(parts) < 17: parts.append("-")

        return {
            "CSM Name": parts[0], "Customer Name": parts[1], "Lead Intent": parts[2],
            "⚠️ Brand Trust": parts[3], "📝 Brand Detail": parts[4],
            "⚠️ Value": parts[5], "📝 Value Detail": parts[6],
            "⚠️ ROI": parts[7], "📝 ROI Detail": parts[8],
            "⚠️ Spiritual": parts[9], "📝 Spiritual Detail": parts[10],
            "⚠️ Technical": parts[11], "📝 Technical Detail": parts[12],
            "Urgency: 6L+2L Offer": parts[13], "Urgency: Price Movement": parts[14],
            "Family/Spouse Objections": parts[15], "Pitch Flow Adherence": parts[16]
        }
    except:
        return None

def generate_csm_summary(model, df, csm_name):
    csm_df = df[df['CSM Name'] == csm_name]
    cols = ['Customer Name', 'Lead Intent', '📝 Brand Detail', '📝 Value Detail', 'Urgency: 6L+2L Offer']
    csv_data = csm_df[cols].to_csv(index=False)

    prompt = f"""
    You are a Sales Trainer. Summarize performance for CSM: {csm_name}.
    Data: {csv_data}
    
    **Style:** "Strengths" and "Areas of Improvement" must use format:
    **Category Name:** Description (citing Customer Names).
    
    **Return JSON:** {{ "CSM Name": "{csm_name}", "Strengths": "...", "Areas of Improvement": "..." }}
    """
    try:
        response = model.generate_content(prompt)
        clean = response.text.replace("```json", "").replace("```", "").strip()
        return json.loads(clean)
    except:
        return {"CSM Name": csm_name, "Strengths": "Error", "Areas of Improvement": "Error"}

def to_excel(df_analysis, df_summary):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_analysis.to_excel(writer, sheet_name='Call Analysis', index=False)
        if df_summary is not None:
            df_summary.to_excel(writer, sheet_name='CSM Summary', index=False)
        
        # Styling
        for sheet in writer.sheets:
            ws = writer.sheets[sheet]
            for col in ws.columns:
                col_letter = col[0].column_letter
                width = 80 if sheet == 'CSM Summary' and col_letter in ['B','C'] else 40
                ws.column_dimensions[col_letter].width = width
                
                # Header Style
                header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                ws[f"{col_letter}1"].fill = header_fill
                ws[f"{col_letter}1"].font = header_font

    return output.getvalue()

# --- Main UI ---
st.title("🦚 Vrindavan Global Sales Auditor")
st.markdown("### Navratna v4.6 (Web Edition)")

with st.sidebar:
    st.header("Configuration")
    api_key = st.text_input("Enter Gemini API Key", type="password")
    uploaded_files = st.file_uploader("Upload Transcripts (PDF)", type=['pdf'], accept_multiple_files=True)
    
    run_button = st.button("🚀 Run Analysis", type="primary", disabled=not (api_key and uploaded_files))

if run_button:
    model = get_gemini_model(api_key)
    if not model:
        st.error("❌ Invalid API Key or No Model Found.")
    else:
        results = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # 1. Analyze Calls
        total = len(uploaded_files)
        for idx, uploaded_file in enumerate(uploaded_files):
            status_text.text(f"Processing {idx+1}/{total}: {uploaded_file.name}")
            text = extract_text_from_pdf(uploaded_file.getvalue())
            
            if text.strip():
                # Retry logic
                for attempt in range(3):
                    data = analyze_call(model, text)
                    if data:
                        data['File Name'] = uploaded_file.name
                        results.append(data)
                        break
                    time.sleep(2) # Backoff
            
            progress_bar.progress((idx + 1) / total)
            time.sleep(1)

        if results:
            st.session_state['analysis_results'] = pd.DataFrame(results)
            
            # 2. CSM Summaries
            status_text.text("🧠 Generating Training Summaries...")
            unique_csms = st.session_state['analysis_results']['CSM Name'].unique()
            summaries = []
            
            summ_prog = st.progress(0)
            for i, csm in enumerate(unique_csms):
                summ_prog.progress((i + 1) / len(unique_csms))
                summary_data = generate_csm_summary(model, st.session_state['analysis_results'], csm)
                summaries.append(summary_data)
                time.sleep(1.5)
            
            st.session_state['csm_summary'] = pd.DataFrame(summaries)
            status_text.text("✅ Analysis Complete!")
            summ_prog.empty()
            progress_bar.empty()
        else:
            st.error("❌ No results could be extracted.")

# --- Display Results ---
if st.session_state['analysis_results'] is not None:
    tab1, tab2 = st.tabs(["📋 Detailed Analysis", "👨‍🏫 CSM Summary"])
    
    with tab1:
        st.dataframe(st.session_state['analysis_results'], use_container_width=True)
        
    with tab2:
        if st.session_state['csm_summary'] is not None:
            st.markdown("### Training Feedback")
            # Custom CSS for wrapping text in tables
            st.markdown(
                """<style>
                div[data-testid="stDataFrame"] div[class*="stDataFrame"] { white-space: pre-wrap; }
                </style>""", unsafe_allow_html=True
            )
            st.dataframe(st.session_state['csm_summary'], use_container_width=True)

    # Export Button
    excel_data = to_excel(st.session_state['analysis_results'], st.session_state['csm_summary'])
    st.download_button(
        label="📥 Download Excel Report",
        data=excel_data,
        file_name="Vrindavan_Web_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
